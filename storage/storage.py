"""
storage/json_storage.py — Save jobs to JSON file.
storage/csv_storage.py  — Save jobs to CSV file.
storage/sqlite_storage.py — Save jobs to SQLite database.
"""

# ─── json_storage.py ──────────────────────────────────────────────────────────
import json
import os
import logging
from models import Job
from config import JSON_OUTPUT_PATH, OUTPUT_DIR

logger = logging.getLogger(__name__)


def save_json(jobs: list[Job], path: str = JSON_OUTPUT_PATH) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    data = [job.to_dict() for job in jobs]
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    logger.info(f"Saved {len(jobs)} jobs → {path}")
    return path


# ─── csv_storage.py ───────────────────────────────────────────────────────────
import csv
from config import CSV_OUTPUT_PATH


def save_csv(jobs: list[Job], path: str = CSV_OUTPUT_PATH) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if not jobs:
        return path

    fields = [
        "id", "title", "company", "location", "salary", "job_type",
        "experience_level", "remote", "skill_match_score", "matched_skills",
        "skills_required", "source_portal", "url", "apply_url",
        "posted_date", "scraped_at", "description",
    ]

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for job in jobs:
            row = job.to_dict()
            row["matched_skills"] = ", ".join(row["matched_skills"])
            row["skills_required"] = ", ".join(row["skills_required"])
            writer.writerow(row)

    logger.info(f"Saved {len(jobs)} jobs → {path}")
    return path


# ─── sqlite_storage.py ────────────────────────────────────────────────────────
import sqlite3
from config import SQLITE_DB_PATH


def save_sqlite(jobs: list[Job], db_path: str = SQLITE_DB_PATH) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS jobs (
            id TEXT PRIMARY KEY,
            title TEXT,
            company TEXT,
            location TEXT,
            salary TEXT,
            job_type TEXT,
            experience_level TEXT,
            remote INTEGER,
            skill_match_score REAL,
            matched_skills TEXT,
            skills_required TEXT,
            source_portal TEXT,
            url TEXT,
            apply_url TEXT,
            posted_date TEXT,
            description TEXT,
            scraped_at TEXT
        )
    """)

    inserted = 0
    skipped = 0
    for job in jobs:
        d = job.to_dict()
        try:
            cursor.execute("""
                INSERT OR IGNORE INTO jobs VALUES (
                    :id, :title, :company, :location, :salary, :job_type,
                    :experience_level, :remote, :skill_match_score,
                    :matched_skills, :skills_required, :source_portal,
                    :url, :apply_url, :posted_date, :description, :scraped_at
                )
            """, {
                **d,
                "remote": int(d["remote"]),
                "matched_skills": ", ".join(d["matched_skills"]),
                "skills_required": ", ".join(d["skills_required"]),
            })
            if cursor.rowcount:
                inserted += 1
            else:
                skipped += 1
        except sqlite3.Error as e:
            logger.warning(f"SQLite insert error: {e}")

    conn.commit()
    conn.close()
    logger.info(f"SQLite: {inserted} inserted, {skipped} skipped (duplicates) → {db_path}")
    return db_path


# ─── excel_storage.py ─────────────────────────────────────────────────────────
from datetime import datetime as _dt

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


def save_excel(jobs: list[Job], path: str | None = None) -> str:
    """Save jobs to Excel with scrape date and time columns."""
    if Workbook is None:
        raise ImportError("openpyxl is required for Excel export. Run: pip install openpyxl")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    now = _dt.now()
    if path is None:
        path = os.path.join(OUTPUT_DIR, f"jobs_{now.strftime('%Y%m%d_%H%M%S')}.xlsx")

    scrape_date = now.strftime("%Y-%m-%d")
    scrape_time = now.strftime("%H:%M:%S")

    fields = [
        "scrape_date", "scrape_time", "scraped_at",
        "id", "title", "company", "location", "salary", "job_type",
        "experience_level", "remote", "skill_match_score", "matched_skills",
        "skills_required", "source_portal", "url", "apply_url",
        "posted_date", "description",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, name in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col, value=name.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    def _excel_value(field: str, value):
        if field == "remote":
            return "Yes" if value else "No"
        if isinstance(value, list):
            return ", ".join(str(v) for v in value)
        if value is None:
            return ""
        return value

    for row_idx, job in enumerate(jobs, 2):
        d = job.to_dict()
        row = {**d, "scrape_date": scrape_date, "scrape_time": scrape_time}
        for col_idx, field in enumerate(fields, 1):
            ws.cell(row=row_idx, column=col_idx, value=_excel_value(field, row.get(field, "")))

    for col_idx, field in enumerate(fields, 1):
        max_len = len(field)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    wb.save(path)
    logger.info(f"Saved {len(jobs)} jobs → {path}")
    return path
